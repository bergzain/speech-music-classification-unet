import logging
from pathlib import Path
from typing import Optional, Tuple, Dict, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.lines import Line2D
import os, re

def setup_logging():
    """Set up logging configuration."""
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

def get_run_metrics(run_path: Path) -> Dict[str, List[float]]:
    """
    Extract metrics from the given run path.

    Parameters:
    - run_path (Path): The path to the run directory.

    Returns:
    - Dict[str, List[float]]: A dictionary with metric names as keys and lists of values.
    """
    metrics = {}
    try:
        metrics_path = run_path / 'metrics'
        if metrics_path.exists():
            for metric_file in metrics_path.iterdir():
                with metric_file.open('r') as f:
                    values = [float(line.strip().split()[1]) for line in f if line.strip()]
                    metrics[metric_file.name] = values
    except Exception as e:
        logging.error(f"Failed to read metrics from {run_path}: {e}")
    return metrics

def get_best_and_least_loss(metrics: Dict[str, List[float]]) -> Tuple[float, float]:
    """
    Get the best accuracy and least loss from the metrics.

    Parameters:
    - metrics (Dict[str, List[float]]): The metrics dictionary.

    Returns:
    - Tuple[float, float]: The best accuracy and least loss.
    """
    best_accuracy = max(metrics.get('val_accuracy', [0]))
    least_loss = min(metrics.get('val_loss', [float('inf')]))
    return best_accuracy, least_loss

def get_run_name(run_path: Path) -> Optional[str]:
    """
    Retrieve the run name from the given run path.

    Parameters:
    - run_path (Path): The path to the run directory.

    Returns:
    - Optional[str]: The run name if found, else None.
    """
    try:
        tags_path = run_path / 'tags'
        run_name_file = tags_path / 'mlflow.runName'
        if run_name_file.exists():
            with run_name_file.open('r') as f:
                run_name = f.read().strip()
            return run_name
    except Exception as e:
        logging.error(f"Failed to read run name from {run_path}: {e}")
    return None

def parse_run_name(run_name: str) -> Optional[Tuple[str, str, int, int]]:
    """
    Parse the run name to extract model details.

    Parameters:
    - run_name (str): The run name string.

    Returns:
    - Optional[Tuple[str, str, int, int]]: Parsed model name, transformation type,
      number of features, and length in seconds. Returns None if parsing fails.
    """
    pattern = r'([A-Za-z0-9_]+)_([A-Za-z\-]+)_([0-9]+)_len([0-9]+)S'
    match = re.match(pattern, run_name)
    if match:
        model_name, transformation_type, num_features, length_in_seconds = match.groups()
        return model_name, transformation_type, int(num_features), int(length_in_seconds)
    else:
        logging.warning(f"Run name does not match expected format: {run_name}")
        return None

def get_model_parameters(model_name: str, model_sizes: Dict[str, int]) -> int:
    """
    Get the number of parameters for the given model.

    Parameters:
    - model_name (str): The name of the model.
    - model_sizes (Dict[str, int]): A dictionary of model sizes.

    Returns:
    - int: Number of parameters. Returns 0 if model not found.
    """
    if model_name in model_sizes:
        return model_sizes[model_name]
    else:
        logging.warning(f"Unknown model name: {model_name}. Defaulting number of parameters to 0.")
        return 0

def highlight_cells(ws, df, max_acc, min_loss):
    """
    Highlight cells in the Excel sheet based on conditions.

    Parameters:
    - ws: The worksheet object.
    - df: The DataFrame containing the data.
    - max_acc (float): The maximum accuracy value.
    - min_loss (float): The minimum loss value.
    """
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    header_font = Font(bold=True)
    
    # Highlight headers
    for cell in ws[1]:
        cell.font = header_font

    # Highlight max accuracy
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            if cell.value == max_acc:
                cell.fill = yellow_fill

    # Highlight min loss
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row:
            if cell.value == min_loss:
                cell.fill = green_fill

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

def process_run(run_path: Path, model_sizes: Dict[str, int]) -> Optional[Tuple]:
    """
    Process a single run directory to extract metrics and run details.

    Parameters:
    - run_path (Path): The path to the run directory.
    - model_sizes (Dict[str, int]): A dictionary of model sizes.

    Returns:
    - Optional[Tuple]: A tuple containing run details and metrics. Returns None if processing fails.
    """
    run_name = get_run_name(run_path)
    if run_name:
        metrics = get_run_metrics(run_path)
        best_accuracy, least_loss = get_best_and_least_loss(metrics)
        if best_accuracy > 0 and least_loss < float('inf'):
            parsed_data = parse_run_name(run_name)
            if parsed_data:
                model_name, transformation_type, num_features, length_in_seconds = parsed_data
                num_parameters = get_model_parameters(model_name, model_sizes)
                return (run_name, model_name, transformation_type, num_features, length_in_seconds, num_parameters, best_accuracy, least_loss)
    return None

def plot_ball_chart(df: pd.DataFrame, output_dir: Path):
    """
    Plot a bubble chart (ball chart) of the model performance.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    if df.empty:
        logging.info("No data available to plot.")
        return

    plt.figure(figsize=(12, 8))
    sns.set(style="whitegrid")

    # Map colors to transformation types
    transformation_types = df['Transformation Type'].unique()
    colors = sns.color_palette("hsv", len(transformation_types))
    color_dict = dict(zip(transformation_types, colors))

    # Map markers to model names
    model_names = df['Model Name'].unique()
    markers = ['o', 's', 'D', '^', 'v', '<', '>', 'p', '*', 'h']
    marker_dict = dict(zip(model_names, markers))

    # Calculate sizes based on number of parameters
    min_size = 50
    max_size = 300  # Adjusted to make bubbles smaller
    sizes = df['Number of Parameters']
    sizes_scaled = (sizes - sizes.min()) / (sizes.max() - sizes.min())
    sizes_scaled = sizes_scaled * (max_size - min_size) + min_size

    # Plot each data point
    for idx, row in df.iterrows():
        x = row['Least Loss']
        y = row['Best Accuracy']
        size = sizes_scaled.iloc[idx]
        color = color_dict[row['Transformation Type']]
        marker = marker_dict[row['Model Name']]
        plt.scatter(x, y, s=size, c=[color], marker=marker, edgecolors='k', alpha=0.7)

    # Create custom legends
    # Legend for markers (Model Names)
    marker_handles = [Line2D([0], [0], marker=marker_dict[model], color='w', label=model,
                             markerfacecolor='gray', markersize=8, markeredgecolor='k') for model in model_names]
    
    # Legend for colors (Transformation Types)
    color_handles = [Line2D([0], [0], marker='o', color='w', label=trans_type,
                            markerfacecolor=color_dict[trans_type], markersize=8, markeredgecolor='k') for trans_type in transformation_types]
    
    # Combine legends
    first_legend = plt.legend(handles=marker_handles, title='Model Names', loc='upper right', bbox_to_anchor=(1.15, 1))
    plt.gca().add_artist(first_legend)
    plt.legend(handles=color_handles, title='Transformation Types', loc='upper right', bbox_to_anchor=(1.15, 0.6))

    plt.xlabel('Least Loss', fontsize=14)
    plt.ylabel('Best Accuracy', fontsize=14)
    plt.title('Model Performance Bubble Chart', fontsize=16)
    plt.grid(True)
    plt.tight_layout()

    # Save the plot
    output_dir.mkdir(exist_ok=True)
    plot_path = output_dir / 'model_performance_ball_chart.png'
    plt.savefig(plot_path, dpi=300, bbox_inches='tight')
    logging.info(f"Bubble chart saved to {plot_path}")
    plt.show()
    
    
def plot_best_performing_models(df: pd.DataFrame, output_dir: Path):
    """
    Plot the best performing models by accuracy.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    top_models = df.nlargest(5, 'Best Accuracy')

    plt.figure(figsize=(10, 6))
    sns.barplot(x='Run Name', y='Best Accuracy', data=top_models)
    plt.title('Top 5 Best Performing Models by Accuracy')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot
    plot_path = output_dir / 'best_performing_models.png'
    plt.savefig(plot_path, dpi=300)
    plt.show()

def plot_effect_of_features(df: pd.DataFrame, output_dir: Path):
    """
    Plot the effect of different features on accuracy.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    plt.figure(figsize=(10, 6))
    sns.boxplot(x='Transformation Type', y='Best Accuracy', data=df)
    plt.title('Effect of Different Features on Accuracy')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot
    plot_path = output_dir / 'effect_of_features.png'
    plt.savefig(plot_path, dpi=300)
    plt.show()

def plot_effect_of_chunk_length(df: pd.DataFrame, output_dir: Path):
    """
    Plot the effect of chunk length on model performance.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    plt.figure(figsize=(10, 6))
    sns.boxplot(x='Length of Chunks (s)', y='Best Accuracy', data=df)
    plt.title('Effect of Chunk Length on Accuracy')
    plt.tight_layout()

    # Save the plot
    plot_path = output_dir / 'effect_of_chunk_length.png'
    plt.savefig(plot_path, dpi=300)
    plt.show()

def plot_number_of_features_vs_performance(df: pd.DataFrame, output_dir: Path):
    """
    Plot the relationship between number of features and performance.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    plt.figure(figsize=(10, 6))
    sns.scatterplot(x='Number of Features', y='Best Accuracy', hue='Transformation Type', style='Model Name', s=100, data=df)
    plt.title('Number of Features vs. Accuracy')
    plt.tight_layout()

    # Save the plot
    plot_path = output_dir / 'number_of_features_vs_performance.png'
    plt.savefig(plot_path, dpi=300)
    plt.show()

def plot_model_architecture_comparison(df: pd.DataFrame, output_dir: Path):
    """
    Plot comparison of model architectures by performance.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    plt.figure(figsize=(10, 6))
    sns.boxplot(x='Model Name', y='Best Accuracy', data=df)
    plt.title('Comparison of Model Architectures')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    # Save the plot
    plot_path = output_dir / 'model_architecture_comparison.png'
    plt.savefig(plot_path, dpi=300)
    plt.show()

def plot_effect_of_delta_and_delta_delta(df: pd.DataFrame, output_dir: Path):
    """
    Plot the effect of delta and delta-delta features on model performance.

    Parameters:
    - df (pd.DataFrame): The DataFrame containing the data.
    - output_dir (Path): The directory to save the plot.
    """
    delta_df = df[df['Transformation Type'].str.contains('delta')]

    plt.figure(figsize=(10, 6))
    sns.boxplot(x='Transformation Type', y='Best Accuracy', data=delta_df)
    plt.title('Effect of Delta and Delta-Delta Features on Accuracy')
    plt.tight_layout()

    # Save the plot
    plot_path = output_dir / 'effect_of_delta_and_delta_delta.png'
    plt.savefig(plot_path, dpi=300)
    plt.show()


def main():
    """
    Main function to process MLflow runs, generate Excel report, and create visualizations.
    """
    setup_logging()

    # Define the MLflow directory and output directory
    # You can modify these paths as needed
    mlflow_dir = Path("/Users/zainhazzouri/projects/Master-thesis-experiments/mlflow/")  # Replace with your MLflow directory path
    output_dir = Path("/Users/zainhazzouri/projects/Bachelor_Thesis/results")  # Replace with your desired output directory

    # Alternatively, prompt the user for input
    # mlflow_dir = Path(input("Enter the path to the MLflow directory: ").strip())
    # output_dir = Path(input("Enter the path to the output directory (default 'output'): ").strip() or "output")

    # Define model sizes
    model_sizes = {
        "U_Net": 34526084,
        "AttentionUNet": 34877486,
        "R2U_Net": 39091460,
        "R2AttU_Net": 39442992
    }

    runs = []

    # Iterate over runs
    for root, dirs, _ in os.walk(mlflow_dir):
        for dir_name in dirs:
            run_path = Path(root) / dir_name
            if (run_path / 'metrics').is_dir():
                run_data = process_run(run_path, model_sizes)
                if run_data:
                    runs.append(run_data)

    if not runs:
        logging.info("No runs found or no valid data extracted.")
        return

    # Create DataFrame
    columns = ['Run Name', 'Model Name', 'Transformation Type', 'Number of Features', 'Length of Chunks (s)',
               'Number of Parameters', 'Best Accuracy', 'Least Loss']
    df = pd.DataFrame(runs, columns=columns)

    # Specify data types
    df = df.astype({
        'Number of Features': 'int32',
        'Length of Chunks (s)': 'int32',
        'Number of Parameters': 'int64',
        'Best Accuracy': 'float32',
        'Least Loss': 'float32'
    })

    # Save DataFrame to Excel
    excel_file = output_dir / 'run_metrics.xlsx'
    output_dir.mkdir(exist_ok=True)
    df.to_excel(excel_file, index=False)

    # Load workbook and apply highlighting
    wb = load_workbook(excel_file)
    ws = wb.active

    max_acc = df['Best Accuracy'].max()
    min_loss = df['Least Loss'].min()
    highlight_cells(ws, df, max_acc, min_loss)

    # Save the workbook
    wb.save(excel_file)
    logging.info(f"Metrics saved to {excel_file}")

    # Plot the ball chart
    plot_ball_chart(df, output_dir)
    
    
    # Generate plots for findings
    plot_best_performing_models(df, output_dir)
    plot_effect_of_features(df, output_dir)
    plot_effect_of_chunk_length(df, output_dir)
    plot_number_of_features_vs_performance(df, output_dir)
    plot_model_architecture_comparison(df, output_dir)
    plot_effect_of_delta_and_delta_delta(df, output_dir)


if __name__ == "__main__":
    main()