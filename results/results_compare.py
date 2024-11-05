import logging
from pathlib import Path
from typing import Optional, Tuple, Dict, List
import yaml
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.lines import Line2D
import os
import re
from matplotlib.patches import Patch
from datetime import datetime
import numpy as np
from brokenaxes import brokenaxes
from sklearn.utils import resample

def setup_logging():
    """Set up logging configuration."""
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

def get_run_timestamp(run_path: Path) -> Optional[str]:
    """
    Get the timestamp of when the experiment was run from meta.yaml file.
    
    Parameters:
    - run_path (Path): Path to the run directory
    
    Returns:
    - Optional[str]: Timestamp string if found, None otherwise
    """
    try:
        meta_file = run_path / 'meta.yaml'
        if meta_file.exists():
            with open(meta_file, 'r') as f:
                meta_data = yaml.safe_load(f)
                start_time = meta_data.get('start_time', None)
                if start_time:
                    dt = datetime.fromtimestamp(start_time / 1000.0)
                    return dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        logging.error(f"Failed to read timestamp from {run_path}: {e}")
    return None

def get_run_metrics(run_path: Path) -> Dict[str, List[float]]:
    """Extract metrics from the given run path."""
    metrics = {}
    try:
        metrics_path = run_path / 'metrics'
        if metrics_path.exists():
            for metric_file in metrics_path.iterdir():
                with metric_file.open('r') as f:
                    values = [float(line.strip().split()[1]) for line in f if line.strip()]
                    metrics[metric_file.name] = values
    except Exception as e:
        logging.error(f"Failed to read metrics from {run_path}: {e}")
    return metrics

def get_best_and_least_loss(metrics: Dict[str, List[float]]) -> Tuple[float, float]:
    """Get the best accuracy and least loss from the metrics."""
    best_accuracy = max(metrics.get('val_accuracy', [0]))
    least_loss = min(metrics.get('val_loss', [float('inf')]))
    return best_accuracy, least_loss

def get_run_name(run_path: Path) -> Optional[str]:
    """Retrieve the run name from the given run path."""
    try:
        tags_path = run_path / 'tags'
        run_name_file = tags_path / 'mlflow.runName'
        if run_name_file.exists():
            with run_name_file.open('r') as f:
                run_name = f.read().strip()
            return run_name
    except Exception as e:
        logging.error(f"Failed to read run name from {run_path}: {e}")
    return None

def parse_run_name(run_name: str) -> Optional[Tuple[str, str, int, float]]:
    """Parse the run name to extract model details."""
    pattern = r'([A-Za-z0-9_]+)_([A-Za-z\-]+)_([0-9]+)_len([0-9]*\.?[0-9]+)S'
    match = re.match(pattern, run_name)
    if match:
        model_name, transformation_type, num_features, length_in_seconds = match.groups()
        return model_name, transformation_type, int(num_features), float(length_in_seconds)
    else:
        logging.warning(f"Run name does not match expected format: {run_name}")
        return None

def get_model_parameters(model_name: str, model_sizes: Dict[str, int]) -> int:
    """Get the number of parameters for the given model."""
    if model_name in model_sizes:
        return model_sizes[model_name]
    else:
        logging.warning(f"Unknown model name: {model_name}. Defaulting number of parameters to 0.")
        return 0

def highlight_cells(ws, df, max_acc, min_loss):
    """Highlight cells in the Excel sheet based on conditions."""
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    header_font = Font(bold=True)
    
    # Highlight headers
    for cell in ws[1]:
        cell.font = header_font

    # Highlight max accuracy
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            if cell.value == max_acc:
                cell.fill = yellow_fill

    # Highlight min loss
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):
        for cell in row:
            if cell.value == min_loss:
                cell.fill = green_fill

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

def process_run(run_path: Path, model_sizes: Dict[str, int]) -> Optional[Tuple]:
    """Process a single run directory to extract metrics and run details."""
    run_name = get_run_name(run_path)
    if run_name:
        metrics = get_run_metrics(run_path)
        best_accuracy, least_loss = get_best_and_least_loss(metrics)
        timestamp = get_run_timestamp(run_path)
        
        if best_accuracy > 0 and least_loss < float('inf'):
            parsed_data = parse_run_name(run_name)
            if parsed_data:
                model_name, transformation_type, num_features, length_in_seconds = parsed_data
                num_parameters = get_model_parameters(model_name, model_sizes)
                return (run_name, model_name, transformation_type, num_features, 
                       length_in_seconds, num_parameters, best_accuracy, least_loss, timestamp)
    return None

def plot_ball_chart(df: pd.DataFrame, output_dir: Path):
    """Plot a bubble chart of the model performance."""
    plt.figure(figsize=(12, 8))
    sns.set(style="whitegrid")

    # Map colors to transformation types
    transformation_types = df['Transformation Type'].unique()
    colors = sns.color_palette("hsv", len(transformation_types))
    color_dict = dict(zip(transformation_types, colors))

    # Map markers to model names
    model_names = df['Model Name'].unique()
    markers = ['o', 's', 'D', '^', 'v', '<', '>', 'p', '*', 'h']
    marker_dict = dict(zip(model_names, markers))

    # Calculate sizes based on number of parameters
    min_size = 50
    max_size = 300
    sizes = df['Number of Parameters']
    sizes_scaled = (sizes - sizes.min()) / (sizes.max() - sizes.min())
    sizes_scaled = sizes_scaled * (max_size - min_size) + min_size

    # Plot each data point
    for idx, row in df.iterrows():
        x = row['Least Loss']
        y = row['Best Accuracy']
        size = sizes_scaled.iloc[idx]
        color = color_dict[row['Transformation Type']]
        marker = marker_dict[row['Model Name']]
        plt.scatter(x, y, s=size, c=[color], marker=marker, edgecolors='k', alpha=0.7)

    # Create custom legends
    marker_handles = [Line2D([0], [0], marker=marker_dict[model], color='w', label=model,
                           markerfacecolor='gray', markersize=8, markeredgecolor='k') 
                     for model in model_names]
    
    color_handles = [Line2D([0], [0], marker='o', color='w', label=trans_type,
                           markerfacecolor=color_dict[trans_type], markersize=8, markeredgecolor='k') 
                    for trans_type in transformation_types]

    first_legend = plt.legend(handles=marker_handles, title='Model Names', 
                            loc='upper right', bbox_to_anchor=(1.15, 1))
    plt.gca().add_artist(first_legend)
    plt.legend(handles=color_handles, title='Transformation Types', 
              loc='upper right', bbox_to_anchor=(1.15, 0.6))

    plt.xlabel('Least Loss', fontsize=14)
    plt.ylabel('Best Accuracy', fontsize=14)
    plt.title('Model Performance Bubble Chart', fontsize=16)
    plt.grid(True)
    plt.tight_layout()

    plot_path = output_dir / 'model_performance_ball_chart.png'
    plt.savefig(plot_path, dpi=300, bbox_inches='tight')
    plt.close()

def plot_best_performing_models(df: pd.DataFrame, output_dir: Path):
    """Plot the best performing models by accuracy."""
    top_models = df.nlargest(5, 'Best Accuracy')

    plt.figure(figsize=(10, 6))
    sns.barplot(x='Run Name', y='Best Accuracy', data=top_models)
    plt.title('Top 5 Best Performing Models by Accuracy')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    plot_path = output_dir / 'best_performing_models.png'
    plt.savefig(plot_path, dpi=300)
    plt.close()

def plot_effect_of_features(df: pd.DataFrame, output_dir: Path):
    """Plot the effect of different features on accuracy."""
    plt.figure(figsize=(10, 6))
    sns.boxplot(x='Transformation Type', y='Best Accuracy', data=df)
    plt.title('Effect of Different Features on Accuracy')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    plot_path = output_dir / 'effect_of_features.png'
    plt.savefig(plot_path, dpi=300)
    plt.close()

def plot_effect_of_chunk_length(df: pd.DataFrame, output_dir: Path, model_name: str = "U_Net"):
    """Plot the effect of chunk length on model performance."""
    model_df = df[df['Model Name'] == model_name]
    
    if model_df.empty:
        logging.info(f"No data available for model '{model_name}'")
        return

    plt.figure(figsize=(10, 6))
    sns.boxplot(x='Length of Chunks (s)', y='Best Accuracy', data=model_df)
    plt.title(f'Effect of Chunk Length on Accuracy for {model_name}')
    plt.xlabel('Length of Chunks (s)')
    plt.ylabel('Best Accuracy')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()

    plot_path = output_dir / f'effect_of_chunk_length_{model_name}.png'
    plt.savefig(plot_path, dpi=300)
    plt.close()


def plot_ball_chart_5_sec(df: pd.DataFrame, output_dir: Path):
    """Plot a bubble chart of the model performance for 5-second experiments."""
    # Filter for 5-second experiments and reset index
    df_5sec = df[df['Length of Chunks (s)'] == 5.0].reset_index(drop=True)
    
    if df_5sec.empty:
        logging.info("No 5-second models available to plot.")
        return
    
    plt.figure(figsize=(12, 8))
    sns.set(style="whitegrid")

    # Map colors to transformation types
    transformation_types = df_5sec['Transformation Type'].unique()
    colors = sns.color_palette("hsv", len(transformation_types))
    color_dict = dict(zip(transformation_types, colors))

    # Map markers to model names
    model_names = df_5sec['Model Name'].unique()
    markers = ['o', 's', 'D', '^']  # One marker for each model type
    marker_dict = dict(zip(model_names, markers))

    # Model sizes dictionary
    model_sizes = {
        "U_Net": 34526084,
        "AttentionUNet": 34877486,
        "R2U_Net": 39091460,
        "R2AttU_Net": 39442992
    }

    # Calculate sizes based on number of parameters
    min_size = 50
    max_size = 300
    sizes = df_5sec['Number of Parameters']
    sizes_scaled = (sizes - sizes.min()) / (sizes.max() - sizes.min())
    sizes_scaled = sizes_scaled * (max_size - min_size) + min_size

    # Plot each data point
    for idx, row in df_5sec.iterrows():
        x = row['Least Loss']
        y = row['Best Accuracy']
        size = sizes_scaled[idx]
        color = color_dict[row['Transformation Type']]
        marker = marker_dict[row['Model Name']]
        plt.scatter(x, y, s=size, c=[color], marker=marker, edgecolors='k', alpha=0.7)

    # Create markers legend
    marker_handles = [Line2D([0], [0], 
                           marker=marker_dict[model], 
                           color='w', 
                           label=model,
                           markerfacecolor='gray', 
                           markersize=8, 
                           markeredgecolor='k') 
                     for model in model_names]

    # Create color legend
    color_handles = [Line2D([0], [0], 
                          marker='o', 
                          color='w', 
                          label=trans_type,
                          markerfacecolor=color_dict[trans_type], 
                          markersize=8, 
                          markeredgecolor='k') 
                    for trans_type in transformation_types]

    # Create size legend based on model architectures
    size_handles = []
    shape_order = {
        'o': ('U_Net', 34.5),
        's': ('AttentionUNet', 34.9),
        'D': ('R2U_Net', 39.1),
        '^': ('R2AttU_Net', 39.4)
    }

    for marker, (model, param_count) in shape_order.items():
        size = min_size + (max_size - min_size) * (model_sizes[model] - min(model_sizes.values())) / (max(model_sizes.values()) - min(model_sizes.values()))
        size_handles.append(
            Line2D([0], [0],
                  marker=marker,
                  color='w',
                  label=f'{param_count}M params',
                  markerfacecolor='gray',
                  markersize=np.sqrt(size/np.pi) * 0.5,
                  markeredgecolor='k')
        )
    # Add legends
    first_legend = plt.legend(handles=marker_handles, 
                            title='Model Architecture',
                            loc='upper right', 
                            bbox_to_anchor=(1.15, 1))
    plt.gca().add_artist(first_legend)
    
    second_legend = plt.legend(handles=color_handles, 
                             title='Feature Type',
                             loc='upper right', 
                             bbox_to_anchor=(1.15, 0.7))
    plt.gca().add_artist(second_legend)
    
    plt.legend(handles=size_handles, 
              title='Model Parameters',
              loc='upper right', 
              bbox_to_anchor=(1.15, 0.4))

    plt.xlabel('Least Loss', fontsize=14)
    plt.ylabel('Best Accuracy', fontsize=14)
    plt.title('Model Performance Comparison (5-Second Length)', fontsize=16)
    plt.grid(True)
    
    # Adjust layout to make room for legends
    plt.subplots_adjust(right=0.85)
    
    plot_path = output_dir / 'model_performance_ball_chart_5_sec.png'
    plt.savefig(plot_path, dpi=300, bbox_inches='tight')
    plt.close()

def plot_accuracy_vs_complexity(df: pd.DataFrame, output_dir: Path):
    """Plot accuracy versus model complexity."""
    df_5sec = df[df['Length of Chunks (s)'] == 5.0]
    if df_5sec.empty:
        logging.info("No 5-second models available to plot.")
        return

    plt.figure(figsize=(14, 8))
    sns.set(style="whitegrid")

    transformation_types = df_5sec['Transformation Type'].unique()
    fill_colors = sns.color_palette("hsv", len(transformation_types))
    fill_color_dict = dict(zip(transformation_types, fill_colors))

    model_names = df_5sec['Model Name'].unique()
    ring_colors = sns.color_palette("Set1", len(model_names))
    ring_color_dict = dict(zip(model_names, ring_colors))

    plt.xscale('log')

    for idx, row in df_5sec.iterrows():
        x_val = row['Number of Parameters']
        y_val = row['Best Accuracy']
        model = row['Model Name']
        transformation = row['Transformation Type']

        edge_color = ring_color_dict[model]
        face_color = fill_color_dict[transformation]

        plt.scatter(x_val, y_val, s=200,
                   facecolors=face_color,
                   edgecolors=edge_color,
                   linewidths=2,
                   alpha=0.7)

    ring_handles = [Line2D([0], [0], marker='o', color='w', 
                          label=model,
                          markerfacecolor='white',
                          markeredgecolor=ring_color_dict[model],
                          markersize=10,
                          linewidth=2)
                   for model in model_names]

    fill_handles = [Patch(facecolor=fill_color_dict[trans_type],
                         edgecolor='gray',
                         label=trans_type)
                   for trans_type in transformation_types]

    plt.legend(handles=ring_handles,
              title='Model Names (Ring Colors)',
              loc='upper left',
              bbox_to_anchor=(1.02, 1),
              borderaxespad=0)

    plt.gca().add_artist(plt.legend(handles=fill_handles,
                                  title='Transformation Types (Fill Colors)',
                                  loc='upper left',
                                  bbox_to_anchor=(1.02, 0.6),
                                  borderaxespad=0))

    plt.xlabel('Number of Parameters (log scale)', fontsize=14)
    plt.ylabel('Best Accuracy', fontsize=14)
    plt.title('Accuracy vs. Model Complexity (5-Second Models)', fontsize=16)

    plt.tight_layout(rect=[0, 0, 0.75, 1])
    
    plot_path = output_dir / 'accuracy_vs_complexity.png'
    plt.savefig(plot_path, dpi=300, bbox_inches='tight')
    plt.close()




def main():
    """Main function to process MLflow runs and generate reports."""
    setup_logging()

    mlflow_dir = Path("/Users/zainhazzouri/projects/Master-thesis-experiments/mlflow/")
    output_dir = Path("/Users/zainhazzouri/projects/Bachelor_Thesis/results")

    model_sizes = {
        "U_Net": 34526084,
        "AttentionUNet": 34877486,
        "R2U_Net": 39091460,
        "R2AttU_Net": 39442992
    }

    runs = []
    
    # Process all runs
    for root, dirs, _ in os.walk(mlflow_dir):
        for dir_name in dirs:
            run_path = Path(root) / dir_name
            if (run_path / 'metrics').is_dir():
                run_data = process_run(run_path, model_sizes)
                if run_data:
                    runs.append(run_data)

    if not runs:
        logging.info("No runs found or no valid data extracted.")
        return

    # Create DataFrame
    columns = ['Run Name', 'Model Name', 'Transformation Type', 'Number of Features', 
              'Length of Chunks (s)', 'Number of Parameters', 'Best Accuracy', 
              'Least Loss', 'Timestamp']
    df = pd.DataFrame(runs, columns=columns)

    # Set data types
    df = df.astype({
        'Number of Features': 'int32',
        'Length of Chunks (s)': 'float32',
        'Number of Parameters': 'int64',
        'Best Accuracy': 'float32',
        'Least Loss': 'float32'
    })
    
    # Process timestamps
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    df = df.sort_values('Timestamp')

    # Save to Excel
    output_dir.mkdir(exist_ok=True)
    excel_file = output_dir / 'run_metrics.xlsx'
    df.to_excel(excel_file, index=False)

    # Apply Excel formatting
    wb = load_workbook(excel_file)
    ws = wb.active
    highlight_cells(ws, df, df['Best Accuracy'].max(), df['Least Loss'].min())
    wb.save(excel_file)

    # Generate all plots
    plot_ball_chart(df, output_dir)
    plot_accuracy_vs_complexity(df, output_dir)
    plot_best_performing_models(df, output_dir)
    plot_effect_of_features(df, output_dir)
    plot_effect_of_chunk_length(df, output_dir, model_name="U_Net")
    plot_ball_chart_5_sec(df, output_dir)
    

    # Print summary statistics
    print("\nExperiment Timeline Summary:")
    print(f"First experiment: {df['Timestamp'].min()}")
    print(f"Last experiment: {df['Timestamp'].max()}")
    print(f"Total duration: {df['Timestamp'].max() - df['Timestamp'].min()}")
    print(f"Total number of experiments: {len(df)}")

if __name__ == "__main__":
    main()